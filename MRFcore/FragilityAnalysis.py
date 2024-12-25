import os
import sys
import json
from pathlib import Path
from typing import Literal

import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl as px
from matplotlib.axes import Axes
from matplotlib.figure import Figure
from scipy.stats import norm
from scipy.interpolate import interp1d
from loguru import logger
from MRFcore.get_SSF import get_SFF
if __name__ == "__main__":
    sys.path.append(str(Path(__file__).parent.parent))

"""
最后更新：
2024-04-07: 优化画图代码，增加保存结果图像文件
2024-05-31: 功能增强，可同时处理多种工程需求参数类型，代码结构优化
2024-11-10：增加风险评估功能
2024-12-13: 代码结构优化
"""

logger.remove()
logger.add(
    sink=sys.stdout,
    format="<green>{time:YYYY-MM-DD HH:mm:ss}</green> <red>|</red> <level>{level}</level> <red>|</red> <level>{message}</level>",
    level="DEBUG"
)


def get_x(x: list, y: list, y0: float) -> float:
    """获得横线y=y0与给定曲线的交点横坐标

    Args:
        x (list): 输入曲线的横坐标序列\n
        y (list): 输入曲线的纵坐标序列\n
        y0 (float): 横直线y = y0\n

    Returns:
        float: 曲线与横线交点横坐标
    """
    if y0 < min(y):
        raise ValueError(f'【Error】y0 < min(y) ({y0} < {min(y)})')
    if y0 > max(y):
        raise ValueError(f'【Error】y0 > max(y) ({y0} > {max(y)})')
    for i in range(len(y) - 1):
        if y[i] == y0:
            x0 = x[i]
            return x0
        elif y[i] < y0 <= y[i + 1]:
            k = (y[i + 1] - y[i]) / (x[i + 1] - x[i])
            x0 = x[i] + (y0 - y[i]) / k
            return x0
    else:
        raise ValueError('【Error】未找到交点-1')

def get_y(x: list, y: list, x0: float, error: bool=True, x_start_with: float=None) -> float:
    """获得竖线x=x0与给定曲线的交点纵坐标

    Args:
        x (list): 输入曲线的横坐标序列
        y (list): 输入曲线的纵坐标序列
        x0 (float): 竖直线x = x0
        error (boo, optional): 若x0超出范围，抛出异常or返回None
        x_start_with (float, optional): 从x_start_with开始搜索，默认为None，即从x[0]开始搜索

    Returns:
        float: 曲线与竖线交点纵坐标
    """
    # 获得x=x0与曲线的交点
    if x0 < min(x):
        if error:
            raise ValueError(f'【Error】x0 < min(x) ({x0} < {min(x)})')
        else:
            return None
    if x0 > max(x):
        if error:
            raise ValueError(f'【Error】x0 > max(x) ({x0} > {max(x)})')
        else:
            return None
    x = np.array(x)
    if x_start_with is not None:
        x = x[x >= x_start_with]
    y0 = []  # 所有交点的横坐标
    for i in range(len(x) - 1):
        if x[i] == x0:
            y0i = y[i]
            y0.append(y0i)
        elif x[i] < x0 <= x[i + 1]:
            k = (y[i + 1] - y[i]) / (x[i + 1] - x[i])
            y0i = k * (x0 - x[i]) + y[i]
            y0.append(y0i)  # 如果有多个交点，则取均值
    else:
        if len(y0) == 0:
            raise ValueError('【Error】未找到交点-2')
        return np.mean(y0)

    
def get_percentile_line(
        all_x: list[list],
        all_y: list[list],
        p: float,
        n: int=None,
        x: list | np.ndarray=None
    ) -> tuple[np.ndarray, np.ndarray]:
    """计算IDA曲线簇的百分位线

    Args:
        all_x (list[list]): 所有独立IDA的横坐标
        all_y (list[list]): 所有独立IDA的纵坐标
        p (float): 百分位值
        n (int): 输出的百分位线横坐标的点数量
        x (list | np.ndarray, optional): 百分位线横坐标，默认为None，即自动生成

    Returns:
        tuple[np.ndarray, np.ndarray]: 百分位线的横坐标、纵坐标
    """
    # 计算百分位线
    if x is None:
        x1 = min([min(i) for i in all_x])
        x2 = max([max(i) for i in all_x])
        x = np.linspace(x1, x2, n)  # 百分位线横坐标
    y = []  # 百分位线纵坐标
    for _, xi in enumerate(x):
        # xi: int, yi: list
        yi = []
        for _, (line_x, line_y) in enumerate(zip(all_x, all_y)):
            res = get_y(line_x, line_y, xi, False)
            if res is not None:
                yi.append(res)
        y_percentile = np.percentile(yi, p)
        y.append(y_percentile)
    y = np.array(y)
    return x, y

def get_mean_std_line(all_x: list[list], all_y: list[list], n: int) -> tuple[np.ndarray, np.ndarray]:
    """计算IDA曲线簇的标准差线

    Args:
        all_x (list[list]): 所有独立IDA的横坐标
        all_y (list[list]): 所有独立IDA的纵坐标
        n (int): 输出的STD线横坐标的点数量

    Returns:
        tuple[np.ndarray, np.ndarray]: STD线的横坐标、纵坐标
    """
    # 计算百分位线
    x1 = min([min(i) for i in all_x])
    x2 = max([max(i) for i in all_x])
    x = np.linspace(x1, x2, n)  # 百分位线横坐标
    y_mean = []  # 均值线纵坐标
    y_std = []  # 标准差线纵坐标
    for _, xi in enumerate(x):
        # xi: int, yi: list
        yi = []
        for _, (line_x, line_y) in enumerate(zip(all_x, all_y)):
            res = get_y(line_x, line_y, xi, False)
            if res is not None:
                yi.append(res)
        y_mean_ = np.mean(yi)
        y_std_ = np.std(yi)
        y_mean.append(y_mean_)
        y_std.append(y_std_)
    y_mean = np.array(y_mean)
    y_std = np.array(y_std)
    return x, y_mean, y_std

class FragilityAnalysis():
    available_EDP_types = ['IM', 'IDR', 'DCF', 'PFV', 'PFA', 'RIDR', 'RoofIDR' ,'Shear', 'beamHinge', 'colHinge', 'panelZone']  # 允许的DM类型

    def __init__(self, root: str | Path, EDP_types: list[str],
                 collapse_limit: float=0.1, additional_items: list[str]=None):
        """地震易损性、倒塌易损性分析

        Args:
            root (str | Path): 待读取数据的文件夹的路径
            EDP_types (list[str]): 工程需求参数类型
            * [IDR] - 层间位移角
            * [DCF] - 层间变形集中系数
            * [PFV] - 层间速度
            * [PFA] - 楼层绝对加速度
            * [RIDR] - 残余层间位移角
            * [RoofIDR] - 屋顶层间位移角
            * [Shear] - 层间剪力
            * [beamHinge] - 最大梁铰变形
            * [colHinge] - 最大柱铰变形
            * [panelZone] - 最大节点域变形
            * additional_item - 其他项\n
            collapse_limit (float, optional): 倒塌极限层间位移角，如果给定DM对应的层间位移角角大于给定值，
            则不统计，可设为0.1或0.15，默认为0.1\n
            additional_items (list[str], optional): 其他的要读取的项
        """
        for EDP_type in EDP_types:
            if not EDP_type in self.available_EDP_types:
                raise ValueError(f'`EDP_type`不是可用的类型')
        self.Calc_collapse = False  # 是否有进行倒塌易损性计算
        self.Calc_p = False  # 是否有进行EDP超越概率计算
        self.root = Path(root)
        self.EDP_types = EDP_types
        self.collapse_limit = collapse_limit
        self.additional_items = additional_items
        if self.additional_items:
            self.EDP_types += self.additional_items
        self.has_risk_data = False  # 是否进行了风险评估
        self._init_variables()
        self._check_file()
        self._read_file()
        # ResultFolder = model + '_new'
        # self.init_ReadFile(ResultFolder, EDP_type)

    def _init_variables(self):
        """初始化参数"""
        # calc_IDA
        self.IM_scatter: dict[str, list] = {}  # IDA曲线散点(包含临界倒塌点)
        self.DM_scatter: dict[str, list] = {}
        self.IM_scatter1: dict[str, list] = {}  # IDA曲线散点(不包含临界倒塌点)
        self.DM_scatter1: dict[str, list] = {}
        self.IM_lines: dict[str, list[list]] = {}  # IDA曲线
        self.DM_lines: dict[str, list[list]] = {}
        self.pct_x, self.pct_16, self.pct_50, self.pct_84, self.mean, self.std = {}, {}, {}, {}, {}, {}
        # frag_curve
        self.DM_limits: dict[str, float] = {}  # 各DM的最大值，超过该值则不统计
        self.AB: dict[str, tuple[float, float]] = {}  # ln(DM) = A + B * ln(IM)
        self.R2: dict[str, float] = {}  # R方
        self.ln_IM_line, self.ln_DM_line = {}, {}  # ln(IM)和ln(DM)的拟合曲线
        self.x_frag: dict[str, np.ndarray] = {}  # 易损性曲线横坐标
        self.y_frag: dict[str, list[np.ndarray]] = {}  # 易损性曲线纵坐标(多个)
        self.DS: dict[str, dict[str, float]] = {}  # 损伤状态及对应标签
        self.beta: dict[str, float] = {}  # 不确定性系数
        self.info: dict[str, str] = {}  # 记录拟合参数
        # exceedance_probability
        self.DM_has_fixed_beta: dict[str, float] = {}  # 指定了固定不确定性beta_TOT的DM类型
        self.exceed_mean: dict[str, float]  = {}  # 超越概率均值
        self.exceed_std: dict[str, float]  = {}  # 超越概率标准差
        self.exceed_pct50: dict[str, float] = {}  # 中位值
        self.exceed_x, self.exceed_y = {}, {}  # 超越概率点的横纵坐标
        self.exceed_x_fit, self.exceed_y_fit = {}, {}  # 超越概率点的拟合曲线横纵坐标
        self.exceed_y_fixedBeta = {}  # 采用固定不确定性导致的标准差的易损性曲线
        self.DM_values: dict[str, float] = {}  # 超越概率计算所需的DM值
        # manual_probability
        self.risk_figs: dict[str, Figure] = {}
        self.risk_text: dict[str, str] = {}
        self.risk_EDP_hazard_curves: dict[str, np.ndarray] = {}
        # collapse_evaluation
        self.all_CMR: dict[str, float] = {}  # CMR
        # 可视化
        self.all_figures_1: dict[str, Figure] = {}  # 所有易损性相关图
        self.all_figures_2: dict[str, Figure] = {}  # 所有概率风险相关图
        for EDP_type in self.EDP_types[::-1]:
            fig, exes = plt.subplots(nrows=1, ncols=3, figsize=(16, 5))
            fig.suptitle(f'Risk Analysis of `{EDP_type}`')
            self.all_figures_2[EDP_type] = fig
            plt.tight_layout(w_pad=3)
        for EDP_type in self.EDP_types[::-1]:
            fig, exes = plt.subplots(nrows=2, ncols=2, figsize=(16, 12))
            fig.suptitle(f'Fragility Analysis of `{EDP_type}`')
            self.all_figures_1[EDP_type] = fig
            plt.tight_layout(pad=3)

    def _check_file(self):
        # 数据初步检查
        list_ = ['ground_motions', 'Nstory', 'Nbay', 'running_case']
        for file in list_:
            if not Path.exists(self.root/f'{file}.dat'):
                raise ValueError('【Error】未找到{}'.format(self.root/f'{file}.dat'))
        self.GM_names = np.loadtxt(self.root/'ground_motions.dat', dtype=str, ndmin=1).tolist()  # 地震动名
        self.GM_N = len(self.GM_names)
        self.Nstory = int(np.loadtxt(self.root/'Nstory.dat'))  # 楼层数
        self.Nbay = int(np.loadtxt(self.root/'Nbay.dat'))  # 跨数
        with open(self.root/'notes.dat', 'r') as f:
            self.notes = f.read()
        self.running_case = str(np.loadtxt(self.root / 'running_case.dat', dtype=str))
        if self.running_case != 'IDA':
            raise ValueError('【Error】本程序仅支持处理IDA结果数据')
        logger.success('通过数据文件检查')

    def _read_file(self):
        """读取结果文件"""
        # 读取数据
        self.data: list[pd.DataFrame] = []  # 包含所有时程结果的最值
        columns = self.available_EDP_types
        if self.additional_items:
            columns += self.additional_items
        for idx_gm in range(self.GM_N):
            # 遍历地震动
            gm_name = self.GM_names[idx_gm]
            print(f'  正在读取: {gm_name}    \r', end='')
            num =  1
            df = pd.DataFrame(columns=columns)
            while True:
                # 遍历每个动力增量
                folder = self.root / f'{gm_name}_{num}'
                if not Path.exists(folder):
                    break
                line = []
                IM = np.loadtxt(folder/'Sa.out')  # 地震动强度指标
                IM = round(float(IM), 6)
                line.append(IM)
                if (folder/'层间位移角.out').exists():
                    IDR = np.loadtxt(folder/'层间位移角.out')  # 层间位移角
                    line.append(np.max(np.abs(IDR)))
                else:
                    line.append(0)
                if (folder/'DCF.out').exists():
                    DCF = np.loadtxt(folder/'DCF.out')  # 层间变形集中系数
                    line.append(np.max(np.abs(DCF)))
                else:
                    line.append(0)
                if (folder/'层速度.out').exists():
                    PFV = np.loadtxt(folder/'层速度.out')  # 层间相对速度
                    line.append(np.max(np.abs(PFV)))
                else:
                    line.append(0)
                if (folder/'层加速度(g).out').exists():
                    PFA = np.loadtxt(folder/'层加速度(g).out')  # 层间相对加速度
                    line.append(np.max(np.abs(PFA)))
                else:
                    line.append(0)
                if (folder/'残余层间位移角.out').exists():
                    RIDR = np.loadtxt(folder/'残余层间位移角.out')  # 残余层间位移角
                    line.append(np.max(np.abs(RIDR)))
                else:
                    line.append(0)
                if (folder/'屋顶层间位移角.out').exists():
                    RoofIDR = np.loadtxt(folder/'屋顶层间位移角.out')  # 屋顶层间位移角
                    line.append(np.max(np.abs(RoofIDR)))
                else:
                    line.append(0)
                if (folder/'楼层剪力(kN).out').exists():
                    Shear = np.loadtxt(folder/'楼层剪力(kN).out')  # 层间剪力
                    line.append(np.max(np.abs(Shear)))
                else:
                    line.append(0)
                if (folder/'梁铰变形.out').exists():
                    beamHinge = np.loadtxt(folder/'梁铰变形.out')  # 梁铰变形
                    line.append(np.max(np.abs(beamHinge)))
                else:
                    line.append(0)
                if (folder/'柱铰变形.out').exists():
                    colHinge = np.loadtxt(folder/'柱铰变形.out')  # 柱铰变形
                    line.append(np.max(np.abs(colHinge)))
                else:
                    line.append(0)
                if (folder/'节点域变形.out').exists():
                    panelZone = np.loadtxt(folder/'节点域变形.out')  # 节点域变形
                    line.append(np.max(np.abs(panelZone)))
                else:
                    line.append(0)
                if self.additional_items:
                    for item in self.additional_items:
                        if not (folder/f'{item}.out').exists():
                            raise FileExistsError(f'无法找到文件:', str((folder/f'{item}.out').absolute()))
                        data = np.loadtxt(folder/f'{item}.out')  # 其他的要读取的项
                        line.append(np.max(np.abs(data)))
                df.loc[len(df)] = line
                num += 1
            df = df.sort_values(by='IM')  # 每条地震波按IM列的大小排序
            df.index = [i for i in range(1, len(df) + 1)]  # 重新指定index
            for idx, line in df.iterrows():
                if line['IDR'] >= self.collapse_limit:
                    break
            df = df[df.index<=idx]  # 保留临界倒塌点，删除之后的数据
            self.data.append(df)
        logger.success('已读取数据')

    def calc_IDA(self, EDP_type: str, DM_limit: float=None, marked_idx: int=None):
        """计算IDA曲线

        Args:
            EDP_type (str): 工程需求参数类型
            DM_limit (float, optional): DM的上限值，超过该值则不统计，默认为None
            marked_idx (int, optional): 标记的IDA曲线的索引用于画图时高亮显示，默认为None
        """
        print(f'  正在计算`{EDP_type}`类型的IDA曲线      \r', end='')
        IM_scatter, DM_scatter = [], []  # 所有IDA曲线的散点坐标
        IM_scatter1, DM_scatter1 = [], []  # 所有IDA曲线的散点坐标（不包含倒塌点）
        IM_lines, DM_lines = [], []  # IDA曲线簇(多条)
        if DM_limit is not None:
            self.DM_limits[EDP_type] = DM_limit
        for df in self.data:
            IM_i = df['IM'].to_numpy()
            DM_i = df[EDP_type].to_numpy()
            if (DM_limit is not None) and (np.max(DM_i) > DM_limit):
                # 剔除超过DM_limit的点
                IM_i = np.append(IM_i[DM_i<=DM_limit], IM_i[DM_i>DM_limit][0])
                DM_i = np.append(DM_i[DM_i<=DM_limit], DM_i[DM_i>DM_limit][0])
            IM_i, DM_i = list(IM_i), list(DM_i)
            IM_scatter += IM_i  # 包含临界倒塌点
            DM_scatter += DM_i
            IM_scatter1 += IM_i[:-1]  # 不包含临界倒塌点
            DM_scatter1 += DM_i[:-1]
            IM_lines.append([0] + IM_i)
            DM_lines.append([0] + DM_i)
            pct_x, pct_16 = get_percentile_line(DM_lines, IM_lines, p=16, n=300)
            pct_x, pct_50 = get_percentile_line(DM_lines, IM_lines, p=50, n=300)
            pct_x, pct_84 = get_percentile_line(DM_lines, IM_lines, p=84, n=300)
            _, mean, std = get_mean_std_line(DM_lines, IM_lines, n=300)
        self.IM_scatter[EDP_type] = IM_scatter
        self.DM_scatter[EDP_type] = DM_scatter
        self.IM_scatter1[EDP_type] = IM_scatter1
        self.DM_scatter1[EDP_type] = DM_scatter1
        self.IM_lines[EDP_type] = IM_lines
        self.DM_lines[EDP_type] = DM_lines
        self.pct_x[EDP_type] = pct_x
        self.pct_16[EDP_type] = pct_16
        self.pct_50[EDP_type] = pct_50
        self.pct_84[EDP_type] = pct_84
        self.mean[EDP_type] = mean
        self.std[EDP_type] = std
        # 画图 IDA曲线
        fig = self.all_figures_1[EDP_type]
        ax: Axes = fig.get_axes()[0]
        label1 = 'Uncollapsed points'
        label2 = 'Collapsed points'
        for i, (x, y) in enumerate(zip(self.DM_lines[EDP_type], self.IM_lines[EDP_type])):
            if marked_idx == i:
                ax.plot(x, y, color='blue', markersize=6, label=f'idx={i}', zorder=9999)
            else:
                ax.plot(x, y, color='#BFBFBF', markersize=4)
            ax.scatter(x[:-1], y[:-1], color='#E59EDD', zorder=8999, s=7, label=label1)  # 未倒塌点
            ax.scatter(x[-1], y[-1], color='red', zorder=8999, s=10, label=label2)  # 临界倒塌点
            label1 = None
            label2 = None
        ax.plot(self.pct_x[EDP_type], self.pct_16[EDP_type], label='16%', color='green', linewidth=3, linestyle='--')
        ax.plot(self.pct_x[EDP_type], self.pct_50[EDP_type], label='50%', color='green', linewidth=3)
        ax.plot(self.pct_x[EDP_type], self.pct_84[EDP_type], label='84%', color='green', linewidth=3, linestyle='--')
        ax.set_title(f'IDA curves ({EDP_type})')
        ax.legend()
        if EDP_type == 'IDR':
            ax.set_xlim(0, self.collapse_limit)
        else:
            ax.set_xlim(0)
        if EDP_type in self.DM_limits.keys():
            ax.set_xlim(right=self.DM_limits[EDP_type])
        ax.set_ylim(0)
        ax.set_xlabel('DM')
        ax.set_ylabel('IM')
        logger.success('已计算IDA曲线')

    def frag_curve(self, EDP_type: str, DS: dict[str, float], beta: float, IM_limit: float=None):
        """对概率需求模型进行拟合，计算易损性曲线

        Args:
            EDP_type (str): 工程需求参数名称
            DS (dict[str, float]): 损伤状态标签及对应数值
            beta (float): 表征不确定性的标准差
            IM_limit (float, optional): 易损性曲线的最大横坐标范围，默认为None，取倒塌时84%分为地震动强度的1.8倍
        """
        if not EDP_type in self.EDP_types:
            raise KeyError(f'尚未指定`{EDP_type}`类型，请在`__init__`方法的`EDP_types`参数中添加')
        self.beta[EDP_type] = beta
        self.DS[EDP_type] = DS
        # EDP_type: 损伤指标名称, DS: 损伤状态值
        print(f'  正在计算`{EDP_type}`类型的易损性曲线和概率需求模型      \r', end='')
        IM = np.array(self.IM_scatter1[EDP_type])
        DM = np.array(self.DM_scatter1[EDP_type])
        # 地震易损性曲线的横坐标最大范围(取为84分位线的最大IM值的3倍)
        if IM_limit is None:
            IM2_for_frag_curve = get_y(self.pct_x[EDP_type], self.pct_84[EDP_type], max(DS.values())) * 1.8
        else:
            IM2_for_frag_curve = IM_limit
        # 概率需求模型曲线的横坐标最大范围(取为最大IM值)
        ln_IM, ln_DM = np.log(IM), np.log(DM)
        ln_IM_line = np.linspace(min(ln_IM), max(ln_IM), 1001)
        mean_ln_IM = np.mean(ln_IM)
        mean_ln_DM = np.mean(ln_DM)
        B = np.sum((ln_IM - mean_ln_IM) * (ln_DM - mean_ln_DM)) / np.sum((ln_IM - mean_ln_IM) ** 2)
        A = mean_ln_DM - B * mean_ln_IM
        ln_DM_pre = A + B * ln_IM
        SSR = np.sum((ln_DM - ln_DM_pre) ** 2)
        SST = np.sum((ln_DM - np.mean(ln_DM_pre)) ** 2)
        R2 = 1 - (SSR / SST)
        x_frag = np.linspace(0.001, IM2_for_frag_curve, 1000)  # 易损性曲线x轴
        y_frags: list[np.ndarray] = []  # 易损性曲线y轴
        for _, DS_i in DS.items():
            y_frags.append(norm.cdf((A + B * np.log(x_frag) - np.log(DS_i)) / beta, 0, 1))  # 易损性曲线y轴
        # DM和IM的对数关系
        ln_DM_line = A + B * ln_IM_line
        self.AB[EDP_type] = (A, B)
        self.R2[EDP_type] = R2
        self.ln_IM_line[EDP_type] = ln_IM_line
        self.ln_DM_line[EDP_type] = ln_DM_line
        self.x_frag[EDP_type] = x_frag
        self.y_frag[EDP_type] = y_frags
        self.info[EDP_type] = f'类型`{EDP_type}`\n\n概率模型需求参数(ln(DM) = A + B * ln(IM))：\nA = {A:.6f}\nB = {B:.6f}\nR2 = {R2:.6f}\n\n'
        # 画图: 概率需求曲线
        fig: Figure = self.all_figures_1[EDP_type]
        axes = fig.get_axes()
        ax: Axes = axes[0]
        x = np.array(self.pct_x[EDP_type])
        y = (x / np.exp(A)) ** (1 / B)
        ax.plot(x, y, color='red', label='PSDM')
        ax.legend()
        # 画图: PSDM拟合
        ax: Axes = axes[1]
        ax.plot(ln_IM, ln_DM, 'o')  # 散点
        ax.plot(ln_IM_line, ln_DM_line, 'red', label=f'ln(DM) = {A:.4f} + {B:.4f} * ln(IM)')
        ax.legend()
        ax.set_title('ln(DM) - ln(IM)')
        ax.set_xlabel('ln(IM)')
        ax.set_ylabel('ln(DM)')
        # 画图: 易损性曲线
        ax: Axes = axes[2]
        for i, (label, DS_i) in enumerate(DS.items()):
            ax.plot(x_frag, y_frags[i], label=f'{label} ({DS_i})')
        ax.legend()
        ax.set_title('Fragility curves')
        ax.set_xlim(0)
        ax.set_ylim(0, 1)
        ax.set_xlabel('IM')
        ax.set_ylabel('Exceeding probability')
        logger.success('已完成易损性函数计算和概率需求模型的拟合')

    def exceedance_probability(self,
            EDP_type: str,
            DM_value: float,
            beta_total: dict[str, float | None]=None
        ):
        """计算所有损伤指标超越某值的概率

        Args:
            EDP_type (str): 工程需求参数名称
            DM_values (float): 损伤指标超越DM_value的概率
            beta_total (dict[str, float], optional): 指定确定性来绘制易损性曲线，默认None，即仅考虑计算得到的地震动记录不确定性
        """
        if not EDP_type in self.EDP_types:
            raise KeyError(f'尚未指定`{EDP_type}`类型，请在`__init__`方法的`EDP_types`参数中添加')
        # EDP_type: 损伤指标名称, DM_value: 损伤值
        # 以直线x=DM_value切割所有IDA曲线，获得交点
        IM_lines, DM_lines = self.IM_lines[EDP_type], self.DM_lines[EDP_type]  # 所有IDA曲线
        if DM_value < min([min(i) for i in DM_lines]):
            raise ValueError(f'`{EDP_type}`类型的`{DM_value}`值小于所有IDA曲线的最小DM值({min([max(i) for i in DM_lines])})')
        if DM_value > max([max(i) for i in DM_lines]):
            raise ValueError(f'`{EDP_type}`类型的`{DM_value}`值大于所有IDA曲线的最大DM值({max([max(i) for i in DM_lines])})')
        IM_points = []  # 交点对应的IM纵坐标列表
        for i in range(self.GM_N):
            IM_line, DM_line = IM_lines[i], DM_lines[i]  # 单调IDA曲线
            try:
                y = get_y(DM_line, IM_line, DM_value)
                IM_points.append(y)
            except ValueError:
                logger.warning(f'`{EDP_type}`类型的`{DM_value}`值不在第{i+1}条IDA曲线的DM范围内({min(DM_line)}, {max(DM_line)})')
        if len(IM_points) < self.GM_N:
            logger.error(f'将不进行`{EDP_type}`类型的超越概率统计')
            return
        exceed_mean, exceed_std, exceed_pct50 = np.mean(IM_points), np.std(IM_points), np.percentile(IM_points, 50)  # 均值、标准差、中位值
        exceed_x = sorted(IM_points)  # 超越概率点横坐标(IM)
        exceed_y = np.array([i/self.GM_N for i in range(1, self.GM_N+1)])  # 超越概率点纵坐标(超越概率)
        # 拟合超越概率曲线
        theta = np.median(exceed_x)  # 取中值
        beta = np.std(np.log(exceed_x), ddof=1)  # 对数标准差
        IM1, IM2 = 0.001, max(exceed_x) * 1.2  # 坐标范围
        exceed_x_fit = np.linspace(IM1, IM2, 1001)  # 超越概率曲线横坐标(IM)
        exceed_y_fit = norm.cdf(np.log(exceed_x_fit / theta) / beta, 0, 1)  # 超越概率曲线横坐标(超越概率)
        if beta_total is not None:
            beta_fixed = beta_total  # 采用传入的总不确定性(beta_TOT)
            logger.success(f'为`{EDP_type}`类型指定了体系总不确定性: beta_total = {beta_fixed} (仅用于进行倒塌评估)')
            self.DM_has_fixed_beta[EDP_type] = beta_fixed
        else:
            beta_fixed = beta
        exceed_y_fixedBeta = norm.cdf(np.log(exceed_x_fit / theta) / beta_fixed, 0, 1)
        text = ''  # 统计特征的文本
        text += f'`{EDP_type}`超越{DM_value}的概率特征：\n'
        text += f'均值：{exceed_mean:.6f}\n'
        text += f'标准差：{exceed_std:.6f}\n'
        text += f'中位值：{exceed_pct50:.6f}'
        print(text)
        self.exceed_mean[EDP_type] = exceed_mean
        self.exceed_std[EDP_type] = exceed_std
        self.exceed_pct50[EDP_type] = exceed_pct50
        self.exceed_x[EDP_type] = exceed_x
        self.exceed_y[EDP_type] = exceed_y
        self.exceed_x_fit[EDP_type] = exceed_x_fit
        self.exceed_y_fit[EDP_type] = exceed_y_fit
        self.exceed_y_fixedBeta[EDP_type] = exceed_y_fixedBeta
        self.info[EDP_type] += text
        self.DM_values[EDP_type] = DM_value
        # 画图: 超越概率曲线
        fig: Figure = self.all_figures_1[EDP_type]
        axes = fig.get_axes()
        ax: Axes = axes[3]
        ax.set_title(f'Exceedance probability of {EDP_type}>{DM_value}')
        ax.set_ylabel(f'P({EDP_type}>{DM_value})')
        ax.set_xlabel('IM')
        if EDP_type in self.exceed_x_fit.keys():
            ax.plot(exceed_x_fit, exceed_y_fit, label='Fragility curve')
            ax.plot(exceed_x, exceed_y, 'o', color='red')
            if EDP_type in self.DM_has_fixed_beta.keys():
                ax.plot(exceed_x_fit, exceed_y_fixedBeta, label=f'Fixed beta_TOT = {beta_fixed}')
        ax.set_xlim(0)
        ax.set_ylim(0)
        ax.legend()
        logger.success('已计算超越概率曲线')

    def manual_probability(self,
            EDP_type: str,
            Sa_range: tuple[float, float],
            EDP_range: tuple[float, float],
            hazard_curve: np.ndarray,
            fragility_type: Literal['computed', 'PSDM']='computed',
            density: float=1000
        ):
        """年度超越概率计算

        Args:
            EDP (str): 工程需求参数类型
            Sa_range (tuple[float, float]): 灾害作用范围
            EDP_range (tuple[float, float]): 工程需求参数范围
            hazard_curve (np.ndarray): 灾害曲线(二维数组，两列)
            fragility_type (Literal['computed', 'PSDM']): 使用哪种易损性曲线('computed': 基于IDA数据直接拟合; 'PSDM': 基于概率地震需求模型)
            density (float): 采样点密度
        """
        Sa_min, Sa_max = Sa_range
        EDP_min, EDP_max = EDP_range
        if EDP_min == 0:
            EDP_min = 1e-8
        x_Sa = np.linspace(Sa_min, Sa_max, density)  # Sa轴坐标
        get_log10_harzard_curve = interp1d(np.log10(hazard_curve[:, 0]), np.log10(hazard_curve[:, 1]), kind='cubic', fill_value='extrapolate', bounds_error=False)
        x_temp = np.linspace(hazard_curve[0, 0], hazard_curve[-1, 0], 1000)
        if EDP_type not in self.EDP_types:
            raise KeyError(f'尚未指定`{EDP_type}`类型，请在`__init__`方法的`EDP_types`参数中添加')
        x_EDP = np.linspace(EDP_min, EDP_max, density)  # IDA曲线横坐标(密集点)
        y_ls = np.zeros((self.GM_N, density))  # IDA曲线纵坐标(密集点)
        for i, df in enumerate(self.data):
            x_points = [0] + df[EDP_type].to_list()
            y_points = [0] + df['IM'].to_list()
            _, y = get_percentile_line([x_points], [y_points], p=0.5, x=x_EDP)  # FIXME: 当x_EDP的最大范围超过x_points的时，无法计算百分位线，导致PFA的计算出错
            y_ls[i] = y
        ln_theta = np.median(np.log(y_ls), axis=0)  # 地震动强度中值
        beta = np.std(np.log(y_ls), axis=0)  # 地震动强度标准差
        gama = np.zeros_like(x_Sa)
        log10_HSa = get_log10_harzard_curve(np.log10(x_Sa))  # 灾害曲线的对数
        HSa = np.power(10, log10_HSa)  # 灾害曲线
        diff_HSa = np.append(0, np.diff(HSa))  # 灾害曲线的差分
        for i, (ln_theta_i, beta_i) in enumerate(zip(ln_theta, beta)):
            EDP_value = x_EDP[i]
            # 遍历EDP值
            if fragility_type == 'computed':
                y_cdf = norm.cdf((np.log(x_Sa) - ln_theta_i) / beta_i, 0, 1)  # 易损性函数(累积概率分布曲线)
            elif fragility_type == 'PSDM':
                if not EDP_type in self.EDP_types:
                    raise KeyError(f'尚未指定`{EDP_type}`类型，请在`__init__`方法的`EDP_types`参数中添加')
                A, B = self.AB[EDP_type]
                y_cdf = norm.cdf((A + B * np.log(x_Sa) - np.log(EDP_value)) / self.beta[EDP_type], 0, 1)
            else:
                raise ValueError(f'Wrong fragility_type: {fragility_type}')
            gama_i = np.sum(y_cdf * np.abs(diff_HSa))  # 易损性函数乘以灾害曲线的差分
            gama[i] = gama_i
        P50years = 1 - np.exp(-gama_i * 50)
        text_risk = f'Risk Evaluation:\nManual probability P[{EDP_type} > {round(x_EDP[-1], 2)}] = {gama_i:.3e}\n'
        text_risk += f'50 Years Probability P[{EDP_type} > {round(x_EDP[-1], 2)}] = {P50years:.2%}'
        print(text_risk)
        self.risk_text[EDP_type] = text_risk
        self.risk_EDP_hazard_curves[EDP_type] = np.array([x_EDP, gama]).T
        self.has_risk_data = True
        # 画图: 灾害曲线
        fig = self.all_figures_2[EDP_type]
        axes = fig.get_axes()
        ax: Axes = axes[0]
        ax.loglog(hazard_curve[:, 0], hazard_curve[:, 1], '-o', label='USGS curve')
        ax.loglog(x_temp, np.pow(10, get_log10_harzard_curve(np.log10(x_temp))), label='Cubic Interpolation')
        ax.grid(True)
        ax.set_xlabel('Sa')
        ax.set_ylabel(f'MAF of Sa')
        ax.set_title('Hazard Curve')
        ax.legend()
        ax: Axes = axes[1]
        ax.plot(x_Sa, y_cdf, label=f'P[{EDP_type} > {round(x_EDP[-1], 2)}]')
        ax.set_xlabel('Sa')
        ax.set_ylabel('Probability of Exceedance')
        ax.set_title('Fragility Curve')
        ax.legend()
        ax.grid(True)
        ax: Axes = axes[2]
        ax.semilogy(x_EDP, gama)
        ax.set_xlabel(EDP_type)
        ax.set_ylabel(f'MAF of {EDP_type}')
        ax.set_title(f'{EDP_type} Hazard Curve')
        ax.grid(True)

    def visualization(self):
        """可视化曲线图"""
        plt.show()

    def collapse_evaluation(self,
            T: float,
            MCE_spec: Path | str,
            SF_spec: float=1,
            miuT: float=None,
            SDC: Literal['B', 'C', 'Dmin', 'Dmax', 'other']=None,
        ):
        """Collapse assessment based on FEMA P695 and manual probability

        Args:
            T (float): 结构周期(T=CuTa，非一阶周期)
            MCE_spec (Path | str): MCE规范谱所在文件路径(文件为两列，一列周期一列谱值)
            SF_spec (float, optional): 将规范谱的谱值进行放大，默认放大系数为1
            miuT (float, optional): Period-based ductility，默认None
            SDC (Literal['B', 'C', 'Dmin', 'Dmax', 'other'], optional): Seismic design category，默认None
        """
        MCE_spec = Path(MCE_spec)
        if not MCE_spec.exists():
            raise FileExistsError(f'无法找到文件：{str(MCE_spec.absolute())}')
        if 'IDR' not in self.EDP_types:
            raise ValueError('未进行`IDR`类型的易损性分析')
        spec_data = np.loadtxt(MCE_spec)
        spec_data[:, 1] *= SF_spec
        Sa_MCE = get_y(spec_data[:, 0], spec_data[:, 1], T)  # 一阶周期对应的MCE谱谱值
        SCT_5 = get_x(self.exceed_x_fit['IDR'], self.exceed_y_fixedBeta['IDR'], 0.05)  # 不同倒塌概率对应的倒塌强度
        SCT_10 = get_x(self.exceed_x_fit['IDR'], self.exceed_y_fixedBeta['IDR'], 0.1)
        SCT_15 = get_x(self.exceed_x_fit['IDR'], self.exceed_y_fixedBeta['IDR'], 0.15)
        SCT_20 = get_x(self.exceed_x_fit['IDR'], self.exceed_y_fixedBeta['IDR'], 0.20)
        SCT_25 = get_x(self.exceed_x_fit['IDR'], self.exceed_y_fixedBeta['IDR'], 0.25)
        SCT_50 = get_x(self.exceed_x_fit['IDR'], self.exceed_y_fixedBeta['IDR'], 0.5)
        CMR = SCT_50 / Sa_MCE
        CMR5 = SCT_50 / SCT_5
        CMR10 = SCT_50 / SCT_10
        CMR15 = SCT_50 / SCT_15
        CMR20 = SCT_50 / SCT_20
        CMR25 = SCT_50 / SCT_25
        print(f'SCT_50 = {SCT_50:.3f}, Sa_MCE = {Sa_MCE:.3f}')
        ACMR = 0
        ACMR5 = 0
        ACMR10 = 0
        ACMR15 = 0
        ACMR20 = 0
        ACMR25 = 0
        print(f'CMR = {CMR:.3f}')
        print(f'CMR10 = {CMR10:.3f}')
        print(f'CMR20 = {CMR20:.3f}')
        if miuT is not None:
            ssf = get_SFF(T, miuT, SDC)
            ACMR = CMR * ssf
            ACMR5 = CMR5 * ssf
            ACMR10 = CMR10 * ssf
            ACMR15 = CMR15 * ssf
            ACMR20 = CMR20 * ssf
            ACMR25 = CMR25 * ssf
            print(f'ACMR = {ACMR:.3f}')
            print(f'ACMR10 = {ACMR10:.3f}')
            print(f'ACMR20 = {ACMR20:.3f}')
        self.all_CMR['CMR'] = CMR
        self.all_CMR['CMR5'] = CMR5
        self.all_CMR['CMR10'] = CMR10
        self.all_CMR['CMR15'] = CMR15
        self.all_CMR['CMR20'] = CMR20
        self.all_CMR['CMR25'] = CMR25
        self.all_CMR['ACMR'] = ACMR
        self.all_CMR['ACMR5'] = ACMR5
        self.all_CMR['ACMR10'] = ACMR10
        self.all_CMR['ACMR15'] = ACMR15
        self.all_CMR['ACMR20'] = ACMR20
        self.all_CMR['ACMR25'] = ACMR25

    def save_data(self, output_path: str | Path):
        """保存分析结果
            
            Args:
                output_path (str | Path): 保存结果的文件夹路径
        """
        output_path = Path(output_path)
        if not Path.exists(output_path):
            os.makedirs(output_path)
        print('正在保存数据...\r', end='')
        for EDP_type in self.EDP_types:
            if EDP_type in self.all_figures_1.keys():
                self.all_figures_1[EDP_type].savefig(output_path / f'FragilityAnalysis_{EDP_type}.png', dpi=600)  # 保存曲线图
            if EDP_type in self.all_figures_2.keys():
                self.all_figures_2[EDP_type].savefig(output_path / f'RiskAnalysis_{EDP_type}.png', dpi=600)  # 保存曲线图
            # 1 IDA曲线
            wb = px.Workbook()
            ws1 = wb.active
            ws1.title = 'IDA曲线'
            for i, (x, y) in enumerate(zip(self.DM_lines[EDP_type], self.IM_lines[EDP_type])):
                ws1.cell(1, 2*i+1, i+1)
                ws1.merge_cells(start_row=1, start_column=2*i+1, end_row=1, end_column=2*i+2)
                for j, (xi, yi) in enumerate(zip(x, y)):
                    ws1.cell(j+2, 2*i+1, value=xi)
                    ws1.cell(j+2, 2*i+2, value=yi)
            ws2 = wb.create_sheet('分位线')
            ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            ws2.cell(1, 1, '16%分位线')
            ws2.cell(1, 3, '50%分位线')
            ws2.cell(1, 5, '84%分位线')
            for i, (x, y) in enumerate(zip(self.pct_x[EDP_type], self.pct_16[EDP_type])):
                ws2.cell(i+2, 1, x)
                ws2.cell(i+2, 2, y)
            for i, (x, y) in enumerate(zip(self.pct_x[EDP_type], self.pct_50[EDP_type])):
                ws2.cell(i+2, 3, x)
                ws2.cell(i+2, 4, y)
            for i, (x, y) in enumerate(zip(self.pct_x[EDP_type], self.pct_84[EDP_type])):
                ws2.cell(i+2, 5, x)
                ws2.cell(i+2, 6, y)
            wb.save(output_path / f'IDA曲线_{EDP_type}.xlsx')
            # 2 概率需求模型
            wb = px.Workbook()
            ws = wb.active
            ws.title = 'ln(DM)-ln(IM)'
            ws.cell(1, 1, 'ln(DM)-ln(IM)散点')
            ws.cell(1, 3, '拟合值')
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
            ws.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
            if EDP_type in self.ln_IM_line.keys():
                for i, (x, y) in enumerate(zip(np.log(self.IM_scatter1[EDP_type]), np.log(self.DM_scatter1[EDP_type]))):
                    ws.cell(i+2, 1, x)
                    ws.cell(i+2, 2, y)
                for i, (x, y) in enumerate(zip(self.ln_IM_line[EDP_type], self.ln_DM_line[EDP_type])):
                    ws.cell(i+2, 3, x)
                    ws.cell(i+2, 4, y)
            wb.save(output_path / f'概率需求模型_{EDP_type}.xlsx')
            # 3 地震易损性、超越概率(倒塌易损性)
            wb = px.Workbook()
            ws1 = wb.active
            ws1.title = '地震易损性'
            ws1.cell(1, 1, 'IM')
            ws1.cell(1, 2, '超越概率')
            if EDP_type in self.x_frag.keys():
                ws1.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
                ws1.merge_cells(start_row=1, start_column=2, end_row=1, end_column=1+len(self.DS[EDP_type]))
                for i, label in enumerate(self.DS[EDP_type].keys()):
                    ws1.cell(2, i+2, label)
                for i, x in enumerate(self.x_frag[EDP_type]):
                    ws1.cell(3+i, 1, x)
                for i in range(len(self.y_frag[EDP_type])):
                    for j, y in enumerate(self.y_frag[EDP_type][i]):
                        ws1.cell(3+j, i+2, y)
            if EDP_type in self.exceed_x.keys():
                ws2 = wb.create_sheet(f'超越概率P({EDP_type}>{self.DM_values[EDP_type]})')
                ws2.cell(1, 1, '实际(散点)')
                ws2.cell(1, 3, '拟合(曲线)')
                ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
                ws2.merge_cells(start_row=1, start_column=3, end_row=1, end_column=4)
                for i, (x, y) in enumerate(zip(self.exceed_x[EDP_type], self.exceed_y[EDP_type])):
                    ws2.cell(2+i, 1, x)
                    ws2.cell(2+i, 2, y)
                for i, (x, y) in enumerate(zip(self.exceed_x_fit[EDP_type], self.exceed_y_fit[EDP_type])):
                    ws2.cell(2+i, 3, x)
                    ws2.cell(2+i, 4, y)
                if EDP_type in self.exceed_y_fixedBeta.keys():
                    ws2.cell(1, 5, '拟合(固定beta_TOT)')
                    ws2.merge_cells(start_row=1, start_column=5, end_row=1, end_column=6)
                    for i, (x, y) in enumerate(zip(self.exceed_x_fit[EDP_type], self.exceed_y_fixedBeta[EDP_type])):
                        ws2.cell(2+i, 5, x)
                        ws2.cell(2+i, 6, y)
            wb.save(output_path / f'易损性曲线_{EDP_type}.xlsx')
            # 保存计算结果参数
            if EDP_type in self.info.keys():
                with open(output_path / f'概率特征_{EDP_type}.out', 'w') as f:
                    f.write(self.info[EDP_type])
            with open(output_path / f'CMR.json', 'w') as f:
                json.dump(self.all_CMR, f, indent=4)
        if self.has_risk_data:
            for EDP_type in self.risk_EDP_hazard_curves.keys():
                with open(output_path / f'risk_info_{EDP_type}.out', 'w') as f:
                    f.write(self.risk_text[EDP_type])
                np.savetxt(output_path / f'hazard_curve_{EDP_type}.out', self.risk_EDP_hazard_curves[EDP_type], fmt='%.4e')
        plt.close()
        logger.success('已保存数据')



    